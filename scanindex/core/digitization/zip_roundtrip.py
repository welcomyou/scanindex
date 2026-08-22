"""Round-trip import of an exported archive ZIP back into Step 2 / Kho.

A ZIP produced by the external export path (see `main_window._arc_export_external`)
bundles the final PDFs plus the aggregated `MetaDuLieu.xlsx` under
`HSLTCQ/METADATA/` — and, when the "kèm .json.zst" setting is on (default),
each PDF's canonical `.json.zst` sidecar (OCR text + KIE annotations) next
to it:

    <dossier>.zip
    └── HSLTCQ/
        └── METADATA/
            ├── MetaDuLieu.xlsx            # Hồ sơ + Văn bản sheets
            ├── <stem>-001.pdf
            ├── <stem>-001.pdf.json.zst    # canonical sidecar (optional)
            └── <stem>-NNN.pdf

This module reverses that. Two entry points share one extraction core:

  * `parse_export_zip` — reopen the ZIP into Step 2 (`ArchiveStep2Kie`) for
    editing. PDFs land in a dedicated `_zip_input/` temp folder; each doc's
    `json_path` points at the extracted sidecar when one was bundled, so the
    viewer shows bbox/field overlays immediately. The operator can still
    click "Xử lý" to re-run OCR + KIE from scratch (fresh overlays into
    `_step2_kie/`).
  * `parse_export_zip_for_kho` — feed the ZIP straight into Kho lưu trữ via
    `Importer.import_dossier` **without re-running OCR/KIE**: blocks, KIE
    fields and annotations all come from the bundled sidecars. Docs without
    a sidecar are skipped and reported.

Because the workbook is a lossy, one-way projection, the reconstruction is
*not* byte-for-byte:

  - ZIPs exported with the sidecars disabled (or by older versions) have no
    `.json.zst` companions: bbox highlighting is unavailable on the Step 2
    viewer and the Kho-direct import path has nothing to work with.
  - `chuyen_de`, `chu_thich`, `is_unstructured` are not in the workbook and
    default to empty / False.
  - `ma_dinh_danh` is not stored as a workbook column; it is parsed from the
    ZIP file name (`<MãĐD>-<MãPhông>-<MụcLục>-<HồSơ>.zip`). For the generic
    `HSLTCQ.zip` name we fall back to the other identity codes where possible.

These limitations were confirmed acceptable for the "reopen a ZIP to fix a
few fields, then re-export" workflow.
"""
from __future__ import annotations

import os
import re
import zipfile
from typing import Optional

from scanindex.core.canonical_io import companion_for_pdf, resolve_companion
from scanindex.core.digitization.session import IdentityCodes


# ── Reverse of `runner._FORM_TO_COLUMN` ──────────────────────────────
# The aggregated "Văn bản" sheet column → Step-2 section-1 form key. Used to
# reconstruct `doc["metadata"]` when reading a workbook back in. Note the
# NBSP in "Số của văn\xa0bản" matches the official template verbatim.
_COLUMN_TO_FORM = {
    "Tên cơ quan, tổ chức ban hành văn bản": "co_quan_ban_hanh",
    "Tên loại văn bản":                     "loai_van_ban",
    "Số của văn\xa0bản":                    "so_van_ban",
    "Ký hiệu của văn bản":                  "ky_hieu",
    "Ngày, tháng, năm văn bản":             "ngay_ban_hanh",
    "Trích yếu nội dung":                   "trich_yeu",
    "Ngôn ngữ":                             "ngon_ngu",
    "Người ký":                             "nguoi_ky",
    "Độ mật":                               "do_mat",
    "Tờ số trang số":                       "trang_so",
    "Số thứ tự văn bản trong hồ sơ":        "so_thu_tu",
}

# "Hồ sơ" sheet column → IdentityCodes attribute. The 4 base codes
# (ma_dinh_danh / ma_phong / muc_luc / ho_so) are read from the ZIP file name
# first; the workbook only carries display-friendly forms (ten_phong, …).
_HOSO_COLUMN_TO_IDENTITY = {
    "Tiêu đề hồ sơ":    "title",
    "Thời hạn bảo quản": "thoi_han_bao_quan",
    "Phông":            "ten_phong",
    "Mục lục":          "ten_muc_luc",
    "Nhiệm kỳ":         "nhiem_ky",
    "Tình trạng vật lý": "tinh_trang_vat_ly",
    "Số lượng trang":   "so_luong_trang",
    "Số lượng tờ":      "so_luong_to",
}

# Trailing `-NNN.pdf` ordinal (3 digits, zero-padded) — used to keep document
# order stable across the ZIP listing and the workbook rows.
_PDF_ORDINAL_RE = re.compile(r"-(\d{3})\.pdf$", re.IGNORECASE)

# `<MãĐD>-<MãPhông>-<MụcLục>-<HồSơ>.zip` — the 4 leading dash-separated parts
# of the exported ZIP name. Each part is `[A-Za-z0-9]+` (codes are stripped of
# punctuation by `_arc_export_zip_name`), so a plain dash split is safe as long
# as we only take the first 4 segments. Mục lục / hồ sơ are short, ma_phong may
# be an 8-char hex (unstructured mode).
_ZIP_NAME_PARTS_RE = re.compile(
    r"^(?P<a>[^/\\\-]+)-(?P<b>[^/\\\-]+)-(?P<c>[^/\\\-]+)-(?P<d>[^/\\\-]+)\.zip$",
    re.IGNORECASE,
)


class ZipRoundtripError(Exception):
    """Raised when a file does not look like an archive-export ZIP."""


def _parse_zip_name(zip_path: str) -> dict:
    """Pull the 4 base identity codes from `<MãĐD>-<MãPhông>-<MụcLục>-<HồSơ>.zip`.

    Returns an empty dict for the generic `HSLTCQ.zip` fallback so the caller
    can fill the codes from the workbook's "Đơn vị bảo quản số" column instead.
    """
    name = os.path.basename(zip_path)
    m = _ZIP_NAME_PARTS_RE.match(name)
    if not m:
        return {}
    return {
        "ma_dinh_danh": m.group("a").strip(),
        "ma_phong": m.group("b").strip(),
        "muc_luc": m.group("c").strip(),
        "ho_so": m.group("d").strip(),
    }


def _read_sheet_rows(xlsx_path: str, sheet_name: str) -> list[dict]:
    """Read one sheet of the workbook into a list of `{header: value}` dicts.

    Uses openpyxl read-only mode (same pattern as `repository/importer.py`).
    Empty rows and rows without a "Tên tệp" (Văn bản) value are skipped.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in headers]
        out: list[dict] = []
        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            rec = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                rec[h] = "" if v is None else str(v)
            out.append(rec)
        return out
    finally:
        wb.close()


def _row_to_metadata(row: dict) -> dict:
    """Project one "Văn bản" sheet row onto the Step-2 section-1 form keys.

    `do_mat` empty cell → "Thường" (the UI default; the export path collapses
    "Thường" to blank, so we restore it here so the dropdown shows the right
    value).
    """
    meta = {}
    for col, form_key in _COLUMN_TO_FORM.items():
        v = (row.get(col) or "").strip()
        if form_key == "do_mat" and not v:
            v = "Thường"
        meta[form_key] = v
    return meta


def _count_pdf_pages(pdf_path: str) -> int:
    """Page count of one extracted PDF. Used to seed trang_so when the
    archive didn't carry it. Returns 0 on any failure (caller treats 0 as
    1 page for the running-number computation so ordering stays sane)."""
    if not pdf_path or not os.path.isfile(pdf_path):
        return 0
    try:
        import fitz
        with fitz.open(str(pdf_path)) as f:
            return max(1, int(f.page_count))
    except Exception:
        try:
            from pypdf import PdfReader
            return max(1, len(PdfReader(pdf_path).pages))
        except Exception:
            return 0


def _backfill_trang_so_and_stt(documents: list[dict]) -> None:
    """Seed trang_so / so_thu_tu when every doc's metadata is missing them.

    Archives exported by 1.1.3 and earlier have no "Tờ số trang số" /
    "Số thứ tự văn bản trong hồ sơ" columns, so after `_row_to_metadata`
    every doc has blank trang_so / so_thu_tu. Rather than ship blank cells
    on the next export, derive the running numbering now from the extracted
    PDFs' page counts:

      * so_thu_tu = 1, 2, 3, … (document order)
      * trang_so  = cumulative start page, doc[i+1] starts at
                    trang_so[i] + page_count[i]

    Only runs when ALL docs lack the value — partial state (some docs
    already carry trang_so, e.g. a 1.1.4 archive) is left untouched so we
    never clobber preserved data. Mutates `documents` in place.
    """
    if not documents:
        return
    need_trang = not any(
        str((d.get("metadata", {}) or {}).get("trang_so", "")).strip()
        for d in documents
    )
    need_stt = not any(
        str((d.get("metadata", {}) or {}).get("so_thu_tu", "")).strip()
        for d in documents
    )
    if not (need_trang or need_stt):
        return
    try:
        from scanindex.core.digitization.metadata_export import (
            compute_trang_so, compute_so_thu_tu,
        )
    except Exception:
        return
    page_counts = [
        _count_pdf_pages(d.get("output_path") or d.get("pdf_path") or "")
        for d in documents
    ]
    trang = compute_trang_so(page_counts, first_default=1)
    stt = compute_so_thu_tu(len(documents), first_default=1)
    for i, d in enumerate(documents):
        meta = d.setdefault("metadata", {})
        if need_trang:
            meta["trang_so"] = str(trang[i])
        if need_stt:
            meta["so_thu_tu"] = str(stt[i])


def _identity_from_hoso(identity: IdentityCodes,
                        hoso_rows: list[dict]) -> IdentityCodes:
    """Fill the dossier-level fields (title, retention, fonds name, …) from
    the single "Hồ sơ" sheet row. The 4 base codes are only touched when the
    ZIP name was the generic `HSLTCQ.zip` (i.e. still empty)."""
    if not hoso_rows:
        return identity
    row = hoso_rows[0]
    for col, attr in _HOSO_COLUMN_TO_IDENTITY.items():
        v = (row.get(col) or "").strip()
        if v:
            setattr(identity, attr, v)
    # When the ZIP name carried no codes, "Đơn vị bảo quản số" holds ho_so.
    if not identity.ho_so:
        dv = (row.get("Đơn vị bảo quản số") or "").strip()
        if dv:
            identity.ho_so = dv
    return identity


def _pdf_ordinal(name: str) -> int:
    """Sort key: trailing -NNN of the PDF name, else a large number so
    non-conforming names sort last but stay stable."""
    m = _PDF_ORDINAL_RE.search(name)
    return int(m.group(1)) if m else 999999


def _extract_export_zip(zip_path: str, out_dir: str) -> tuple[list[str], str]:
    """Extract an archive-export ZIP's `HSLTCQ/METADATA/` contents into
    `out_dir`: the workbook (as `_MetaDuLieu.xlsx`), every PDF, and every
    bundled `.json.zst` sidecar (named `<pdf>.json.zst` by the exporter, so
    `companion_for_pdf` finds it next to each extracted PDF).

    Returns `(pdf_names, xlsx_path)`. Only base names are used when writing
    (no zip-slip). Raises `ZipRoundtripError` if the ZIP is not a recognized
    archive export or has no PDFs."""
    if not os.path.isfile(zip_path):
        raise ZipRoundtripError(f"File not found: {zip_path}")

    os.makedirs(out_dir, exist_ok=True)

    pdf_names: list[str] = []
    xlsx_tmp_path: Optional[str] = None
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
            # The export layout uses forward slashes inside the archive.
            metadata_entries = [
                n for n in names
                if n.replace("\\", "/").startswith("HSLTCQ/METADATA/")
            ]
            if not metadata_entries:
                raise ZipRoundtripError(
                    "Missing HSLTCQ/METADATA/ folder — not an archive export ZIP."
                )
            for entry in metadata_entries:
                base = os.path.basename(entry)
                if not base:
                    continue
                low = base.lower()
                if low.endswith(".xlsx") and low.startswith("metadulieu"):
                    target = os.path.join(out_dir, "_MetaDuLieu.xlsx")
                    with zf.open(entry) as src, open(target, "wb") as dst:
                        dst.write(src.read())
                    xlsx_tmp_path = target
                elif low.endswith(".pdf"):
                    target = os.path.join(out_dir, base)
                    with zf.open(entry) as src, open(target, "wb") as dst:
                        dst.write(src.read())
                    pdf_names.append(base)
                elif low.endswith(".json.zst"):
                    # Canonical sidecar bundled by the exporter (optional).
                    target = os.path.join(out_dir, base)
                    with zf.open(entry) as src, open(target, "wb") as dst:
                        dst.write(src.read())
    except ZipRoundtripError:
        raise
    except Exception as e:  # noqa: BLE001 — surface a readable message
        raise ZipRoundtripError(f"Could not read ZIP: {e}") from e

    if not xlsx_tmp_path:
        raise ZipRoundtripError("MetaDuLieu.xlsx not found inside the ZIP.")
    if not pdf_names:
        raise ZipRoundtripError("No PDF documents found inside the ZIP.")

    pdf_names.sort(key=_pdf_ordinal)
    return pdf_names, xlsx_tmp_path


def _identity_and_rows(zip_path: str,
                       xlsx_tmp_path: str) -> tuple[IdentityCodes, list[dict]]:
    """Rebuild the dossier identity (ZIP name + "Hồ sơ" sheet) and the raw
    "Văn bản" sheet rows from an extracted workbook."""
    codes = _parse_zip_name(zip_path)
    identity = IdentityCodes(
        ma_dinh_danh=codes.get("ma_dinh_danh", ""),
        ma_phong=codes.get("ma_phong", ""),
        muc_luc=codes.get("muc_luc", ""),
        ho_so=codes.get("ho_so", ""),
    )
    try:
        hoso_rows = _read_sheet_rows(xlsx_tmp_path, "Hồ sơ")
        identity = _identity_from_hoso(identity, hoso_rows)
    except Exception:
        pass

    vanban_rows: list[dict] = []
    try:
        vanban_rows = _read_sheet_rows(xlsx_tmp_path, "Văn bản")
    except Exception:
        vanban_rows = []
    return identity, vanban_rows


def _row_by_filename(vanban_rows: list[dict]) -> dict:
    """Index Văn bản rows by "Tên tệp" so each PDF can find its metadata
    regardless of listing order inside the workbook."""
    idx = {}
    for r in vanban_rows:
        fn = (r.get("Tên tệp") or "").strip()
        if fn:
            idx[fn] = r
    return idx


def _match_row(name: str, position: int, by_filename: dict,
               vanban_rows: list[dict]) -> dict:
    """Workbook row for the PDF at `position` (0-based). Falls back to
    ordinal matching if the filename column was empty or the names drifted
    (e.g. the operator renamed a file before export)."""
    row = by_filename.get(name)
    if row is None and position < len(vanban_rows):
        row = vanban_rows[position]
    return row or {}


def parse_export_zip(zip_path: str,
                     dest_dir: str) -> tuple[IdentityCodes, list[dict], str]:
    """Unpack an archive-export ZIP and rebuild (identity, documents, input_dir).

    `dest_dir` is the session temp root (e.g. `session.temp_dir()`). PDFs are
    extracted into a dedicated `<dest_dir>/_zip_input/` folder — **not** the
    KIE output dir — so that when the user clicks "Xử lý", the folder-scan
    pipeline re-runs OCR + KIE on them and writes fresh overlays into
    `_step2_kie/`. When the ZIP bundled `.json.zst` sidecars, each doc's
    `json_path` points at the extracted companion so the Step 2 viewer shows
    bbox/field overlays right away (no "Xử lý" needed just to review).
    Returns `(identity, documents, input_dir)` where `input_dir` is the
    `_zip_input/` path the caller should set as the Step 2 input folder.
    Raises `ZipRoundtripError` if the ZIP is not a recognized archive export.
    """
    out_dir = os.path.join(dest_dir, "_zip_input")
    pdf_names, xlsx_tmp_path = _extract_export_zip(zip_path, out_dir)

    identity, vanban_rows = _identity_and_rows(zip_path, xlsx_tmp_path)
    by_filename = _row_by_filename(vanban_rows)

    documents: list[dict] = []
    for i, name in enumerate(pdf_names):
        full_path = os.path.join(out_dir, name)
        meta = _row_to_metadata(_match_row(name, i, by_filename, vanban_rows))
        # Companion bundled by the exporter → viewer overlays work without
        # re-running KIE. Legacy ZIPs (no sidecars) keep json_path empty and
        # behave exactly like before.
        companion = resolve_companion(companion_for_pdf(full_path))
        documents.append({
            "pdf_path": full_path,
            "path": full_path,
            # Point at the extracted PDF so the viewer works before the
            # operator re-runs KIE; once "Xử lý" completes, the pipeline
            # overwrites this with the KIE-overlay output in _step2_kie/.
            "output_path": full_path,
            "ocr_path": None,
            "json_path": str(companion) if companion is not None else "",
            "metadata": meta,
            "zones": {},
            # "Corrected" = reopened from an exported ZIP. Treated as
            # preview-ready (clickable, editable) so the operator can review
            # the parsed metadata before clicking "Xử lý" to re-run KIE.
            "status": "Corrected",
        })

    # Back-compat: archives exported by 1.1.3 (and earlier) lack the
    # "Tờ số trang số" / "Số thứ tự văn bản trong hồ sơ" columns, so the
    # rows above carry blank trang_so / so_thu_tu. Auto-compute them from
    # the extracted PDFs' page counts so the next export always ships a
    # complete workbook, even when the operator reopens a legacy ZIP and
    # hits "Xuất" without clicking a single row first.
    _backfill_trang_so_and_stt(documents)

    return identity, documents, out_dir


def _codes_from_identity(identity: IdentityCodes):
    """Project a reconstructed `IdentityCodes` onto `DossierCodes` for
    `Importer.import_dossier` — mirrors Step 3's "Chuyển vào Kho" mapping.
    Fields the workbook never carried (`chuyen_de`, `chu_thich`,
    `is_unstructured`) default to empty, matching the lossy round-trip
    documented at module level."""
    from scanindex.core.repository.importer import DossierCodes

    return DossierCodes(
        ma_dinh_danh=identity.ma_dinh_danh,
        fonds=identity.ma_phong,
        catalog=identity.muc_luc,
        dossier_code=identity.ho_so,
        fonds_name=getattr(identity, "ten_phong", ""),
        catalog_name=getattr(identity, "ten_muc_luc", ""),
        title=identity.title or f"Hồ sơ {identity.ho_so}",
        is_unstructured=bool(getattr(identity, "is_unstructured", False)),
        retention=identity.thoi_han_bao_quan,
        term=(identity.nhiem_ky or "")[:10],
        storage_unit=identity.ho_so,            # Đơn vị bảo quản số = ho_so
        physical_state=identity.tinh_trang_vat_ly,
        topic=identity.chuyen_de,
        note=identity.chu_thich,
    )


def parse_export_zip_for_kho(zip_path: str,
                             dest_dir: str) -> tuple[object, list[dict], int, str]:
    """Unpack an exported ZIP for direct import into Kho lưu trữ.

    Unlike `parse_export_zip` (Step 2 editing surface, expects a re-run of
    OCR/KIE), this projection feeds `Importer.import_dossier` directly:
    blocks, KIE fields and annotations all come from the bundled `.json.zst`
    sidecars, so **no OCR or KIE is re-run**. Files are extracted into
    `<dest_dir>/_zip_kho/` (a separate folder so it never collides with a
    Step 2 `_zip_input/` reopening of the same ZIP).

    Returns `(codes, documents, skipped_no_companion, out_dir)`:

      * `codes` — `DossierCodes` for `import_dossier` (from the ZIP name +
        the workbook's "Hồ sơ" sheet).
      * `documents` — `[{"pdf_path", "canonical_json_path",
        "target_file_name", "metadata"}]` entries (metadata from the
        "Văn bản" sheet, same keys Step 2 uses).
      * `skipped_no_companion` — PDFs without a bundled sidecar; they are
        *not* included because importing them would require OCR, which this
        path never runs.

    Raises `ZipRoundtripError` if the ZIP is not a recognized archive export.
    """
    out_dir = os.path.join(dest_dir, "_zip_kho")
    pdf_names, xlsx_tmp_path = _extract_export_zip(zip_path, out_dir)

    identity, vanban_rows = _identity_and_rows(zip_path, xlsx_tmp_path)
    codes = _codes_from_identity(identity)
    by_filename = _row_by_filename(vanban_rows)

    documents: list[dict] = []
    skipped_no_companion = 0
    for i, name in enumerate(pdf_names):
        full_path = os.path.join(out_dir, name)
        companion = resolve_companion(companion_for_pdf(full_path))
        if companion is None:
            skipped_no_companion += 1
            continue
        documents.append({
            "pdf_path": full_path,
            "canonical_json_path": str(companion),
            # The export PDF is already named `<identity>-NNN.pdf`; keep it
            # as the in-Kho file name so re-imports are stable.
            "target_file_name": name,
            "metadata": _row_to_metadata(
                _match_row(name, i, by_filename, vanban_rows)
            ),
        })

    return codes, documents, skipped_no_companion, out_dir
