"""Importer for Kho lưu trữ.

Two entry points:

* `import_dossier(identity_codes, documents)` — primary path used by Step 3
  of the archive workflow. Caller hands over already-OCRed PDFs paired with
  their canonical JSON; we read the 14 raw KIE fields straight from the
  JSON's `annotations.field_instances` and persist them lossless.

* `import_folder(source_folder)` — legacy path scanning a folder + an
  HSLTCQ-style metadata.xlsx. xlsx fields are mapped onto the 14 raw KIE
  columns best-effort (lossy: language / copy_type / etc. have no KIE
  equivalent and get dropped).

Per-PDF workflow either path:
1. sha256 dedup inside the target dossier.
2. Copy PDF into repository/pdf/<dossier>/.
3. fitz.get_text("blocks") on the copied PDF → Block list (text overlay
   from prior OCR is preserved by PyMuPDF).
4. chunker.chunk_blocks() → Chunk list.
5. tokenizer.to_no_diacritic + tokenizer.segment for each chunk.
6. SQLite + Tantivy write in 1 logical operation. Status field on
   `documents` and `chunks` lets repair.py recover from a crash.
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import openpyxl

from . import constants as C
from .chunker import Block, Chunk, chunk_blocks, synthesize_metadata_chunk
from .indexer import HybridIndex
from .store import ArchiveStore
from .tokenizer import segment, segment_many, to_no_diacritic


# ---------- Canonical KIE label list (14 fields) ----------
# Order must stay stable: it's the on-disk column order in `documents`.
KIE_LABELS: Tuple[str, ...] = (
    "REGIME_HEADER",
    "ISSUE_ORG_SUPERIOR",
    "ISSUE_ORG_NAME",
    "DOC_NUMBER_SYMBOL",
    "PLACE_DATE",
    "DOC_SUBJECT",
    "ADDRESSEE",
    "RECIPIENTS",
    "SIGNER_ROLE",
    "SIGNER_NAME",
    "URGENCY_MARK",
    "SECRECY_MARK",
    "CIRCULATION_MARK",
    "DOC_TYPE",
)


def _kie_col(label: str) -> str:
    return f"kie_{label.lower()}"


KIE_COLUMNS: Tuple[str, ...] = tuple(_kie_col(l) for l in KIE_LABELS)


# ---------- Header maps from xlsx (legacy import_folder path) ----------

DOC_HEADER_MAP = {
    "Tên cơ quan, tổ chức ban hành văn bản": "issue_org",
    "Tên loại văn bản":         "doc_type",
    "Số của văn bản":           "doc_number",
    "Ký hiệu của văn bản":      "doc_symbol",
    "Ngày, tháng, năm văn bản": "issue_date",
    "Trích yếu nội dung":       "subject",
    "Loại bản":                 "copy_type",
    "Ngôn ngữ":                 "language",
    "Ghi chú":                  "note",
    "Bút tích":                 "inscription",
    "Người ký":                 "signer_name",
    "Chuyên đề":                "topic",
    "Ký hiệu thông tin":        "info_symbol",
    "Từ khóa":                  "keywords",
    "Chế độ sử dụng":           "access_mode",
    "Độ mật":                   "confidentiality",
    "Mức độ tin cậy":           "reliability",
    "Tình trạng vật lý":        "physical_state",
    "Tên tệp":                  "file_name",
    "Thời gian tài liệu":       "doc_period",
}

DOSSIER_HEADER_MAP = {
    "Tiêu đề hồ sơ":               "title",
    "Thời hạn bảo quản":           "retention",
    "Phông":                       "fonds",
    "Mục lục":                     "catalog",
    "Nhiệm kỳ":                    "term",
    "Thời gian bắt đầu":           "start_date",
    "Thời gian kết thúc":          "end_date",
    "Tình trạng vật lý":           "physical_state",
    "Số lượng trang":              "page_count",
    "Tổng số văn bản trong hồ sơ": "doc_count",
    "Độ mật":                      "confidentiality",
    "Đơn vị bảo quản số":          "storage_unit",
    "Thời gian tài liệu":          "doc_period",
}


_DATE_RE = re.compile(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})")
_ISO_DATE_RE = re.compile(r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*$")


def _normalize_date(value) -> Optional[str]:
    if value is None:
        return None
    if hasattr(value, "year"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    iso = _ISO_DATE_RE.match(s)
    if iso:
        y, mo, d = iso.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m = _DATE_RE.search(s)
    if not m:
        return s
    d, mo, y = m.groups()
    if len(y) == 2:
        y = "19" + y if int(y) > 30 else "20" + y
    return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"


def _normalize_step2_display_date(value) -> str:
    """Normalize Step 2 final-form dates to the compact DD/MM/YYYY display."""
    raw = str(value or "").strip()
    if not raw:
        return ""
    iso = _ISO_DATE_RE.match(raw)
    if iso:
        y, mo, d = iso.groups()
        return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"
    m = _DATE_RE.search(raw)
    if not m:
        return raw
    d, mo, y = m.groups()
    if len(y) == 2:
        y = "19" + y if int(y) > 30 else "20" + y
    return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"


def _step2_doc_number(meta: dict) -> str:
    number = " ".join(str(meta.get("so_van_ban") or "").split())
    symbol = " ".join(str(meta.get("ky_hieu") or "").split())
    if number and symbol:
        return f"{number}-{symbol}"
    return number or symbol


def _apply_step2_metadata_overrides(kie_fields: Dict[str, str],
                                    metadata: Optional[dict]) -> Dict[str, str]:
    """Overlay final-form metadata from Step 2 onto raw KIE fields.

    Step 2 lets the user correct the final metadata form without necessarily
    changing the raw KIE bbox annotation. Kho display/search must follow what
    the user accepted in that form.
    """
    if not isinstance(metadata, dict):
        return kie_fields
    out = dict(kie_fields)
    subject = " ".join(str(metadata.get("trich_yeu") or "").split())
    if subject:
        out["kie_doc_subject"] = subject
    doc_number = _step2_doc_number(metadata)
    if doc_number:
        out["kie_doc_number_symbol"] = doc_number
    issue_date = _normalize_step2_display_date(metadata.get("ngay_ban_hanh"))
    if issue_date:
        out["kie_place_date"] = issue_date
    issue_org = " ".join(str(metadata.get("co_quan_ban_hanh") or "").split())
    if issue_org:
        out["kie_issue_org_name"] = issue_org
        out["kie_issue_org_superior"] = ""
    doc_type = " ".join(str(metadata.get("loai_van_ban") or "").split())
    if doc_type:
        out["kie_doc_type"] = doc_type
    signer = " ".join(str(metadata.get("nguoi_ky") or "").split())
    if signer:
        out["kie_signer_name"] = signer
    secrecy = " ".join(str(metadata.get("do_mat") or "").split())
    if secrecy:
        out["kie_secrecy_mark"] = secrecy
    return out


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _allocate_document_instance_id(conn, sha256: str, dossier_id: int) -> str:
    """Return a document-instance id.

    The same PDF bytes may appear in multiple dossiers. Keep the first
    historical id as sha256; suffix later copies so doc_id stays unique while
    sha256 remains available for duplicate/audit checks.
    """
    if conn.execute(
        "SELECT 1 FROM documents WHERE doc_id = ? LIMIT 1", (sha256,)
    ).fetchone() is None:
        return sha256
    suffix = 1
    while True:
        candidate = f"{sha256}-{int(dossier_id)}-{suffix:03d}"
        if conn.execute(
            "SELECT 1 FROM documents WHERE doc_id = ? LIMIT 1", (candidate,)
        ).fetchone() is None:
            return candidate
        suffix += 1


def _normalize_header(value) -> Optional[str]:
    """Collapse whitespace + NBSP so header lookup is robust to template
    drift. The official HSLTCQ template uses NBSP in
    "Số của văn\xa0bản" and a trailing newline in "Tổng số văn bản
    trong hồ sơ\n" in some revisions — both forms must map to the same
    DB field."""
    if value is None:
        return None
    text = str(value).replace("\xa0", " ")
    return " ".join(text.split())


def _read_sheet(ws, header_map: Dict[str, str]) -> List[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    norm_map = {_normalize_header(k): v for k, v in header_map.items()}
    header = [_normalize_header(h) for h in rows[0]]
    out: List[dict] = []
    for r in rows[1:]:
        if not any(v is not None for v in r):
            continue
        d: dict = {}
        for i, h in enumerate(header):
            if h in norm_map and i < len(r):
                d[norm_map[h]] = r[i]
        if d:
            out.append(d)
    return out


def _strip_ocr_suffix(name: str) -> str:
    """`H42-001-01-0123-001_ocr.pdf` → `H42-001-01-0123-001.pdf`."""
    stem, ext = name.rsplit(".", 1) if "." in name else (name, "")
    if stem.endswith("_ocr"):
        stem = stem[:-4]
    return f"{stem}.{ext}" if ext else stem


def _replace_pdf_file(src: Path, dst: Path) -> None:
    """Copy src to dst, replacing any stale PDF already in Kho."""
    src = Path(src).resolve()
    dst = Path(dst).resolve()
    if src == dst:
        return
    dst.parent.mkdir(parents=True, exist_ok=True)
    tmp = dst.with_name(f".{dst.name}.tmp-{int(time.time() * 1000)}")
    try:
        shutil.copy2(src, tmp)
        tmp.replace(dst)
    except Exception as exc:
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass
        raise RuntimeError(
            f"Không thể ghi đè PDF trong Kho: {dst}. "
            "Hãy đóng file PDF nếu đang mở rồi thử lại."
        ) from exc


def _bbox_from_obj(obj: dict) -> Tuple[float, float, float, float]:
    bbox = obj.get("bbox")
    if isinstance(bbox, (list, tuple)) and len(bbox) >= 4:
        try:
            return (
                float(bbox[0]), float(bbox[1]),
                float(bbox[2]), float(bbox[3]),
            )
        except Exception:
            pass
    try:
        x = float(obj.get("x") or 0.0)
        y = float(obj.get("y") or 0.0)
        w = float(obj.get("w") or 0.0)
        h = float(obj.get("h") or 0.0)
        return (x, y, x + w, y + h)
    except Exception:
        return (0.0, 0.0, 0.0, 0.0)


def _page_number(page: dict, fallback_idx: int) -> int:
    try:
        return int(page.get("page_index", fallback_idx)) + 1
    except Exception:
        return fallback_idx + 1


def extract_blocks_from_canonical(canonical: dict) -> List[Block]:
    """Convert OCR canonical JSON into chunker blocks.

    Prefer this over reading the copied PDF text layer: some Vietnamese PDFs
    render correctly but expose mojibake via PyMuPDF, which then pollutes Kho
    search snippets.
    """
    out: List[Block] = []
    pages = canonical.get("pages") or []
    if not isinstance(pages, list):
        return out

    for pi, page in enumerate(pages):
        if not isinstance(page, dict):
            continue
        page_no = _page_number(page, pi)

        raw_blocks = page.get("blocks") or []
        if isinstance(raw_blocks, list) and raw_blocks:
            items = [b for b in raw_blocks if isinstance(b, dict)]
            items.sort(key=lambda b: (
                float((b.get("bbox") or [0, 0, 0, 0])[1] or b.get("y") or 0),
                float((b.get("bbox") or [0, 0, 0, 0])[0] or b.get("x") or 0),
            ))
        else:
            raw_lines = page.get("lines") or []
            items = [ln for ln in raw_lines if isinstance(ln, dict)]
            items.sort(key=lambda ln: (
                int(ln.get("order", 0) or 0),
                float((ln.get("bbox") or [0, 0, 0, 0])[1] or ln.get("y") or 0),
                float((ln.get("bbox") or [0, 0, 0, 0])[0] or ln.get("x") or 0),
            ))

        for idx, item in enumerate(items):
            text = str(item.get("text") or item.get("ocr_text") or "").strip()
            if not text:
                continue
            bbox = _bbox_from_obj(item)
            fs = item.get("font_size")
            if not fs:
                fs = max(1.0, float(bbox[3]) - float(bbox[1]))
            out.append(Block(
                page=page_no,
                block_idx=int(item.get("order", idx) or idx),
                text=text,
                bbox=bbox,
                font_size=float(fs or 12.0),
                region=item.get("region"),
            ))
    return out


def _extract_raw_kie_fields(canonical: dict) -> Dict[str, str]:
    """Pull the 14 KIE labels out of the canonical JSON's annotations
    block, applying the same cleanup Step 2 uses for display so Kho
    columns match what the user saw before pressing 'Chuyển vào Kho'.

    Cleanup pipeline:
    1. `apply_layoutlmv3_schema_postprocess` — schema-level constraints
       (single-line block enforcement, cardinality, signer fragment merge,
       reading-order normalization).
    2. `normalize_subject_for_storage` on DOC_SUBJECT — collapse to a
       single line while preserving the uppercase doc-type prefix.
    3. `single_line_text` on issuing-org fields — DB columns shouldn't
       carry hard line breaks for fields that the form treats as one line.

    The full RAW annotation (pre-cleanup) is preserved separately in
    `kie_annotation_json`, so re-export to other systems can still drop
    back to the unmodified OCR span text if needed."""
    from scanindex.core.kie.text_normalize import (
        normalize_subject_for_storage, single_line_text,
    )
    ann = canonical.get("annotations") or {}

    # Schema-level cleanup — best-effort, falls through with the raw
    # annotation if the postprocess module isn't available (dev box that
    # hasn't pulled the LayoutLMv3 dependencies).
    try:
        from scanindex.core.kie.postprocess import apply_layoutlmv3_schema_postprocess
        ann = apply_layoutlmv3_schema_postprocess(canonical, ann) or ann
    except Exception:
        pass

    fields = ann.get("field_instances") or []
    by_label: Dict[str, List[str]] = {}
    for f in fields:
        label = (f.get("label") or "").strip()
        if not label or label not in KIE_LABELS:
            continue
        text = (f.get("text") or "").strip()
        by_label.setdefault(label, []).append(text)

    doc_type_text = "\n".join(by_label.get("DOC_TYPE", [])).strip()

    out: Dict[str, str] = {}
    for label in KIE_LABELS:
        joined = "\n".join(t for t in by_label.get(label, []) if t).strip()
        if label == "DOC_SUBJECT" and joined:
            joined = normalize_subject_for_storage(joined, doc_type_text)
        elif label in ("ISSUE_ORG_NAME", "ISSUE_ORG_SUPERIOR") and joined:
            joined = single_line_text(joined)
        out[_kie_col(label)] = joined
    return out


def _xlsx_meta_to_kie_fields(doc_meta: dict) -> Dict[str, str]:
    """Best-effort projection from HSLTCQ xlsx columns onto the 14 raw KIE
    fields. Lossy: xlsx columns without KIE counterpart (language, copy_type,
    reliability, ...) are dropped."""
    fields = {col: "" for col in KIE_COLUMNS}
    fields["kie_issue_org_name"] = (doc_meta.get("issue_org") or "").strip()
    fields["kie_doc_subject"]    = (doc_meta.get("subject") or "").strip()
    fields["kie_signer_name"]    = (doc_meta.get("signer_name") or "").strip()
    fields["kie_doc_type"]       = (doc_meta.get("doc_type") or "").strip()
    fields["kie_secrecy_mark"]   = (doc_meta.get("confidentiality") or "").strip()
    fields["kie_circulation_mark"] = (doc_meta.get("access_mode") or "").strip()
    # DOC_NUMBER_SYMBOL: glue the two xlsx columns back together as user typed
    num = (doc_meta.get("doc_number") or "").strip()
    sym = (doc_meta.get("doc_symbol") or "").strip()
    if num and sym:
        fields["kie_doc_number_symbol"] = f"Số: {num}/{sym}"
    elif num:
        fields["kie_doc_number_symbol"] = f"Số: {num}"
    elif sym:
        fields["kie_doc_number_symbol"] = sym
    # PLACE_DATE: just propagate the date (place is not in xlsx)
    iso = _normalize_date(doc_meta.get("issue_date"))
    if iso:
        fields["kie_place_date"] = iso
    return fields


@dataclass
class ImportProgress:
    total: int = 0
    imported: int = 0
    skipped: int = 0
    failed: int = 0
    current_file: str = ""
    message: str = ""


ProgressCallback = Callable[[ImportProgress], None]


# Identity codes from Step 1's IdentityCodes dataclass — passed by value
# so the importer doesn't take a hard dep on archive_session.
@dataclass
class DossierCodes:
    ma_dinh_danh: str
    fonds: str           # mã phông
    catalog: str         # mục lục (≤2 chars)
    dossier_code: str    # hồ sơ (≤5 chars)
    fonds_name: str = ""       # tên phông (≤1000 chars; Excel display)
    catalog_name: str = ""     # tên mục lục (≤1000 chars; Excel display)
    title: str = ""      # tên hồ sơ (≤1000 chars; required when is_unstructured)
    is_unstructured: bool = False
    retention: str = ""        # Thời hạn bảo quản
    term: str = ""             # Nhiệm kỳ
    storage_unit: str = ""     # Đơn vị bảo quản số
    physical_state: str = ""   # Tình trạng vật lý
    topic: str = ""            # Chuyên đề (≤1000 chars)
    note: str = ""             # Chú thích (≤1000 chars)

    @classmethod
    def from_mapping(cls, m: dict) -> "DossierCodes":
        return cls(
            ma_dinh_danh=str(m.get("ma_dinh_danh") or "").strip(),
            fonds=str(m.get("fonds") or m.get("ma_phong") or "").strip(),
            catalog=str(m.get("catalog") or m.get("muc_luc") or "").strip(),
            dossier_code=str(m.get("dossier_code") or m.get("ho_so") or "").strip(),
            fonds_name=str(m.get("fonds_name") or m.get("ten_phong") or "").strip()[:1000],
            catalog_name=str(m.get("catalog_name") or m.get("ten_muc_luc") or "").strip()[:1000],
            title=str(m.get("title") or "").strip()[:1000],
            is_unstructured=bool(m.get("is_unstructured", False)),
            retention=str(m.get("retention") or "").strip(),
            term=str(m.get("term") or "").strip()[:10],
            storage_unit=str(m.get("storage_unit") or "").strip(),
            physical_state=str(m.get("physical_state") or "").strip(),
            topic=str(m.get("topic") or "").strip()[:1000],
            note=str(m.get("note") or "").strip()[:1000],
        )

    def composite_key(self) -> str:
        return f"{self.ma_dinh_danh}-{self.fonds}-{self.catalog}-{self.dossier_code}"


class Importer:
    def __init__(self, store: ArchiveStore, index: HybridIndex):
        self.store = store
        self.index = index

    # ====================================================================
    # PUBLIC — Step 3 path: direct from temp/_step2_kie/ + temp/_step3_signed/
    # ====================================================================

    def import_dossier(self,
                       codes: DossierCodes,
                       documents: List[dict],
                       progress_cb: Optional[ProgressCallback] = None,
                       cancel_check: Optional[Callable[[], bool]] = None,
                       ) -> ImportProgress:
        """Import one dossier whose docs are already OCRed.

        `documents` is a list of `{"pdf_path": str, "canonical_json_path": str}`
        plus optional `metadata` from Step 2's final form. The PDF is the
        final artefact (signed > KIE-overlay) and the JSON is the canonical
        text+annotations file from Step 2. Caller is responsible for picking
        signed-vs-unsigned PDF.
        """
        prog = ImportProgress(total=len(documents))
        if not codes.ma_dinh_danh or not codes.fonds:
            raise ValueError("import_dossier: thiếu mã định danh hoặc mã phông")
        dossier_id = self._upsert_dossier_from_codes(codes)
        import_id = self._begin_import(f"step3:{codes.composite_key()}")
        try:
            self.index.begin_writer()
            for entry in documents:
                if cancel_check and cancel_check():
                    self._end_import(import_id, "cancelled", prog)
                    return prog
                pdf = Path(entry["pdf_path"])
                canonical = Path(entry["canonical_json_path"])
                prog.current_file = pdf.name
                try:
                    inserted = self._import_one_with_canonical(
                        pdf, canonical, codes, dossier_id,
                        target_file_name=entry.get("target_file_name") or "",
                        metadata=entry.get("metadata") or None,
                    )
                    if inserted:
                        prog.imported += 1
                    else:
                        prog.skipped += 1
                except Exception as e:
                    prog.failed += 1
                    prog.message = f"{pdf.name}: {e}"
                if progress_cb:
                    progress_cb(prog)
            self.index.commit()
            self.store.refresh_counters()
            self._end_import(import_id, "completed", prog)
        except Exception as e:
            try:
                self.index.commit()
            except Exception:
                pass
            self._end_import(import_id, "failed", prog, error=str(e))
            raise
        return prog

    # ====================================================================
    # PUBLIC — legacy folder + xlsx import
    # ====================================================================

    def import_folder(self,
                      source_folder: Path,
                      progress_cb: Optional[ProgressCallback] = None,
                      cancel_check: Optional[Callable[[], bool]] = None
                      ) -> ImportProgress:
        source = Path(source_folder)
        xlsx_path = self._find_xlsx(source)
        if xlsx_path is None:
            raise FileNotFoundError(
                f"Không tìm thấy metadata.xlsx trong {source}"
            )
        dossier_rows, doc_rows = self._read_xlsx(xlsx_path)
        dossier_meta = dossier_rows[0] if dossier_rows else {}
        codes = DossierCodes(
            ma_dinh_danh=str(dossier_meta.get("storage_unit") or "xlsx-import"),
            fonds=str(dossier_meta.get("fonds") or "00"),
            catalog=str(dossier_meta.get("catalog") or "00"),
            dossier_code=str(dossier_meta.get("title") or source.name),
            title=str(dossier_meta.get("title") or source.name),
            retention=str(dossier_meta.get("retention") or ""),
            term=str(dossier_meta.get("term") or ""),
            storage_unit=str(dossier_meta.get("storage_unit") or ""),
        )
        dossier_id = self._upsert_dossier_from_codes(codes)

        doc_by_name: Dict[str, dict] = {}
        for d in doc_rows:
            fn = (d.get("file_name") or "").strip()
            if fn:
                doc_by_name[fn.lower()] = d

        pdfs = sorted(p for p in source.glob("*.pdf") if p.is_file())
        prog = ImportProgress(total=len(pdfs))
        import_id = self._begin_import(str(source))

        try:
            self.index.begin_writer()
            for pdf in pdfs:
                if cancel_check and cancel_check():
                    self._end_import(import_id, "cancelled", prog)
                    return prog
                prog.current_file = pdf.name
                doc_meta = doc_by_name.get(pdf.name.lower(), {})
                if not doc_meta:
                    prog.skipped += 1
                    prog.message = f"{pdf.name}: thiếu metadata trong xlsx"
                    if progress_cb:
                        progress_cb(prog)
                    continue
                try:
                    kie_fields = _xlsx_meta_to_kie_fields(doc_meta)
                    inserted = self._import_one_pdf(
                        pdf, kie_fields, kie_annotation_json="",
                        codes=codes, dossier_id=dossier_id,
                        rename_strip_ocr=False,
                    )
                    if inserted:
                        prog.imported += 1
                    else:
                        prog.skipped += 1
                except Exception as e:
                    prog.failed += 1
                    prog.message = f"{pdf.name}: {e}"
                if progress_cb:
                    progress_cb(prog)
            self.index.commit()
            self.store.refresh_counters()
            self._end_import(import_id, "completed", prog)
        except Exception as e:
            try:
                self.index.commit()
            except Exception:
                pass
            self._end_import(import_id, "failed", prog, error=str(e))
            raise

        return prog

    # ====================================================================
    # Internal helpers
    # ====================================================================

    def _find_xlsx(self, folder: Path) -> Optional[Path]:
        for name in ("metadata.xlsx", "MetaDuLieu.xlsx", "Metadata.xlsx"):
            p = folder / name
            if p.exists():
                return p
        candidates = list(folder.glob("*.xlsx"))
        return candidates[0] if candidates else None

    def _read_xlsx(self, path: Path) -> Tuple[List[dict], List[dict]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        dossier_rows: List[dict] = []
        doc_rows: List[dict] = []
        for sheet_name in wb.sheetnames:
            low = sheet_name.lower()
            ws = wb[sheet_name]
            if low.startswith("hồ sơ") or low.startswith("ho so"):
                dossier_rows = _read_sheet(ws, DOSSIER_HEADER_MAP)
            elif low.startswith("văn bản") or low.startswith("van ban"):
                doc_rows = _read_sheet(ws, DOC_HEADER_MAP)
        return dossier_rows, doc_rows

    def _upsert_dossier_from_codes(self, codes: DossierCodes) -> int:
        conn = self.store.connect()
        row = conn.execute(
            "SELECT dossier_id FROM dossiers "
            " WHERE ma_dinh_danh = ? AND fonds = ? "
            "   AND catalog = ? AND dossier_code = ?",
            (codes.ma_dinh_danh, codes.fonds, codes.catalog, codes.dossier_code),
        ).fetchone()
        if row:
            conn.execute(
                "UPDATE dossiers SET fonds_name = ?, catalog_name = ?, "
                "title = COALESCE(NULLIF(?, ''), title), "
                "retention = COALESCE(NULLIF(?, ''), retention), "
                "term = COALESCE(NULLIF(?, ''), term), "
                "storage_unit = COALESCE(NULLIF(?, ''), storage_unit), "
                "physical_state = COALESCE(NULLIF(?, ''), physical_state), "
                "topic = COALESCE(NULLIF(?, ''), topic), "
                "note = COALESCE(NULLIF(?, ''), note), "
                "updated_at = ? WHERE dossier_id = ?",
                (
                    codes.fonds_name or None,
                    codes.catalog_name or None,
                    codes.title or "",
                    codes.retention or "",
                    codes.term or "",
                    codes.storage_unit or "",
                    codes.physical_state or "",
                    codes.topic or "",
                    codes.note or "",
                    int(time.time()),
                    int(row["dossier_id"]),
                ),
            )
            return int(row["dossier_id"])
        now = int(time.time())
        title = codes.title or codes.composite_key()
        cur = conn.execute(
            "INSERT INTO dossiers ("
            "  ma_dinh_danh, fonds, fonds_name, catalog, catalog_name, dossier_code, title,"
            "  is_unstructured, retention, term, storage_unit,"
            "  physical_state, topic, note, created_at"
            ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                codes.ma_dinh_danh, codes.fonds, codes.fonds_name or None,
                codes.catalog, codes.catalog_name or None,
                codes.dossier_code, title,
                1 if codes.is_unstructured else 0,
                codes.retention or None, codes.term or None,
                codes.storage_unit or None,
                codes.physical_state or None,
                codes.topic or None,
                codes.note or None,
                now,
            ),
        )
        return int(cur.lastrowid)

    def _import_one_with_canonical(self,
                                   pdf: Path,
                                   canonical_path: Path,
                                   codes: DossierCodes,
                                   dossier_id: int,
                                   target_file_name: str = "",
                                   metadata: Optional[dict] = None) -> bool:
        canonical = self._load_canonical(canonical_path)
        kie_fields = _extract_raw_kie_fields(canonical)
        kie_fields = _apply_step2_metadata_overrides(kie_fields, metadata)
        ann_block = canonical.get("annotations") or {}
        kie_annotation_json = json.dumps(ann_block, ensure_ascii=False)
        canonical_blocks = extract_blocks_from_canonical(canonical)
        return self._import_one_pdf(
            pdf, kie_fields, kie_annotation_json,
            codes=codes, dossier_id=dossier_id,
            rename_strip_ocr=True,
            target_file_name=target_file_name,
            body_blocks=canonical_blocks,
        )

    @staticmethod
    def _load_canonical(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _import_one_pdf(self,
                        pdf: Path,
                        kie_fields: Dict[str, str],
                        kie_annotation_json: str,
                        *,
                        codes: DossierCodes,
                        dossier_id: int,
                        rename_strip_ocr: bool,
                        target_file_name: str = "",
                        body_blocks: Optional[List[Block]] = None) -> bool:
        sha = _file_sha256(pdf)
        conn = self.store.connect()
        existing_same_dossier = conn.execute(
            "SELECT doc_id FROM documents "
            "WHERE sha256 = ? AND dossier_id = ? AND indexed_status != 'deleted' "
            "LIMIT 1",
            (sha, dossier_id),
        ).fetchone()
        if existing_same_dossier is not None:
            return False
        doc_id = _allocate_document_instance_id(conn, sha, dossier_id)

        # Files inside Kho live under per-dossier subfolder using canonical
        # name (no _ocr suffix). Step-3 path strips it; xlsx path keeps the
        # original filename intact.
        target_name = str(target_file_name or "").strip()
        if not target_name:
            target_name = _strip_ocr_suffix(pdf.name) if rename_strip_ocr else pdf.name
        target_subdir = (
            self.store.archive_path / C.PDF_SUBDIR
            / codes.ma_dinh_danh / codes.fonds / codes.catalog
            / codes.dossier_code
        )
        target_pdf = target_subdir / target_name
        _replace_pdf_file(pdf, target_pdf)

        blocks = list(body_blocks or [])
        if not blocks:
            blocks = self._extract_blocks(target_pdf)
        body_chunks = chunk_blocks(blocks)
        # Per-doc synthesised metadata chunk (KIE summary). Indexed in
        # Tantivy.
        meta_chunk = synthesize_metadata_chunk(kie_fields)
        segment_chunks: List[Chunk] = []
        if meta_chunk is not None:
            segment_chunks.append(meta_chunk)
        segment_chunks.extend(body_chunks)
        segmented_by_obj = {
            id(ch): seg
            for ch, seg in zip(segment_chunks, segment_many(ch.text for ch in segment_chunks))
        }

        with fitz.open(str(target_pdf)) as f:
            page_count = f.page_count

        now = int(time.time())
        self._upsert_document_kie(
            doc_id, dossier_id, target_pdf, target_name,
            kie_fields, kie_annotation_json,
            page_count, status="pending", created_at=now, sha256=sha,
        )

        # ── 1. Metadata chunk (Tantivy only) ──────────────────────────
        if meta_chunk is not None:
            self._insert_metadata_chunk(
                doc_id, dossier_id, meta_chunk, kie_fields, now,
                text_segmented=segmented_by_obj.get(id(meta_chunk)),
            )

        # ── 2. Body chunks (Tantivy) ──
        if body_chunks:
            for ch in body_chunks:
                self._insert_body_chunk_text_only(
                    doc_id, dossier_id, ch, now,
                    text_segmented=segmented_by_obj.get(id(ch)),
                )

        # Mark indexed even if no body chunks (metadata-only doc still
        # searchable via Tantivy on the metadata chunk).
        conn.execute(
            "UPDATE documents SET indexed_status = 'indexed', "
            " indexed_at = ?, updated_at = ? WHERE doc_id = ?",
            (now, now, doc_id),
        )
        conn.execute(
            "UPDATE chunks SET indexed_status = 'indexed' WHERE doc_id = ?",
            (doc_id,),
        )
        return True

    # ---------- Per-chunk insert helpers (v2) ----------

    def _insert_metadata_chunk(self, sha: str, dossier_id: int,
                                ch: Chunk, kie_fields: Dict[str, str],
                                now: int,
                                text_segmented: Optional[str] = None) -> None:
        """Persist the synthesised metadata chunk + index in Tantivy
        with all per-field metadata populated (so per-field boosts hit
        only this row, not every body chunk)."""
        conn = self.store.connect()
        tno = to_no_diacritic(ch.text)
        tseg = text_segmented if text_segmented is not None else segment(ch.text)
        cur = conn.execute(
            "INSERT INTO chunks ("
            "  doc_id, doc_version, chunk_type, page, block_idx,"
            "  text_original, text_no_diacritic, text_segmented,"
            "  bbox, word_count, source_blocks, merge_reason,"
            "  indexed_status, created_at"
            ") VALUES (?, 1, 'metadata', ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)",
            (
                sha, ch.page, ch.block_idx, ch.text, tno, tseg,
                json.dumps(list(ch.bbox)),
                ch.word_count,
                json.dumps(ch.source_blocks),
                ch.merge_reason, now,
            ),
        )
        chunk_id = int(cur.lastrowid)
        # Build the per-Tantivy-field projections from cleaned KIE columns.
        doc_number_proj = kie_fields.get("kie_doc_number_symbol", "")
        signer_proj     = kie_fields.get("kie_signer_name", "")
        issue_org_proj  = " ".join(filter(None, [
            kie_fields.get("kie_issue_org_name", ""),
            kie_fields.get("kie_issue_org_superior", ""),
        ])).strip()
        subject_proj    = kie_fields.get("kie_doc_subject", "")
        recipients_proj = kie_fields.get("kie_recipients", "")
        self.index.add_metadata_chunk(
            doc_id=sha,
            dossier_id=dossier_id,
            chunk_id=chunk_id,
            doc_number=doc_number_proj,
            signer_name=signer_proj,
            issue_org=issue_org_proj,
            subject=subject_proj,
            recipients=recipients_proj,
            metadata_text=ch.text,        # the synthesised KIE prose
        )

    def _insert_body_chunk_text_only(self, sha: str, dossier_id: int,
                                      ch: Chunk, now: int,
                                      text_segmented: Optional[str] = None) -> None:
        """Persist one body chunk + index it in Tantivy."""
        conn = self.store.connect()
        tno = to_no_diacritic(ch.text)
        tseg = text_segmented if text_segmented is not None else segment(ch.text)
        cur = conn.execute(
            "INSERT INTO chunks ("
            "  doc_id, doc_version, chunk_type, page, block_idx,"
            "  text_original, text_no_diacritic, text_segmented,"
            "  bbox, word_count, source_blocks, merge_reason,"
            "  indexed_status, created_at"
            ") VALUES (?, 1, 'body', ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)",
            (
                sha, ch.page, ch.block_idx, ch.text, tno, tseg,
                json.dumps(list(ch.bbox)),
                ch.word_count,
                json.dumps(ch.source_blocks),
                ch.merge_reason, now,
            ),
        )
        chunk_id = int(cur.lastrowid)
        self.index.add_body_text_chunk(
            doc_id=sha,
            dossier_id=dossier_id,
            chunk_id=chunk_id,
            body_original=ch.text,
            body_no_diacritic=tno,
            body_segmented=tseg or "",
        )

    def _extract_blocks(self, pdf_path: Path) -> List[Block]:
        out: List[Block] = []
        with fitz.open(str(pdf_path)) as doc:
            for page_idx, page in enumerate(doc):
                blocks = page.get_text("blocks")
                blocks.sort(key=lambda b: (b[1], b[0]))
                for blk_idx, blk in enumerate(blocks):
                    if len(blk) < 5:
                        continue
                    x0, y0, x1, y1, text = blk[:5]
                    if not text or not text.strip():
                        continue
                    h = float(y1) - float(y0)
                    line_count = max(1, text.count("\n") + 1)
                    fs = h / line_count if line_count else h
                    out.append(Block(
                        page=page_idx + 1,
                        block_idx=blk_idx,
                        text=text.strip(),
                        bbox=(float(x0), float(y0), float(x1), float(y1)),
                        font_size=fs,
                        region=None,
                    ))
        return out

    def _upsert_document_kie(self, doc_id, dossier_id, target_pdf, file_name,
                             kie_fields, kie_annotation_json,
                             page_count, status, created_at,
                             sha256: str = "") -> None:
        rel = str(target_pdf.relative_to(self.store.archive_path)).replace("\\", "/")
        params = {
            "doc_id":              doc_id,
            "dossier_id":          dossier_id,
            "file_name":           file_name,
            "file_path":           rel,
            **{col: kie_fields.get(col, "") for col in KIE_COLUMNS},
            "kie_annotation_json": kie_annotation_json,
            "page_count":          page_count,
            "sha256":              sha256 or doc_id,
            "indexed_status":      status,
            "created_at":          created_at,
        }
        conn = self.store.connect()
        existing = conn.execute(
            "SELECT 1 FROM documents WHERE doc_id = ?", (doc_id,)
        ).fetchone()
        if existing:
            cols = [k for k in params if k not in ("doc_id", "created_at")]
            sets = ", ".join(f"{c} = :{c}" for c in cols)
            conn.execute(
                f"UPDATE documents SET {sets}, updated_at = :now WHERE doc_id = :doc_id",
                {**params, "now": created_at},
            )
        else:
            cols_all = list(params.keys())
            placeholders = ", ".join(f":{c}" for c in cols_all)
            conn.execute(
                f"INSERT INTO documents ({', '.join(cols_all)}) VALUES ({placeholders})",
                params,
            )

    def _begin_import(self, source: str) -> int:
        cur = self.store.connect().execute(
            "INSERT INTO import_history (source_folder, started_at, status) "
            "VALUES (?, ?, 'running')",
            (source, int(time.time())),
        )
        return int(cur.lastrowid)

    def _end_import(self, import_id: int, status: str,
                    prog: ImportProgress, error: Optional[str] = None) -> None:
        self.store.connect().execute(
            "UPDATE import_history SET finished_at = ?, status = ?,"
            " docs_imported = ?, docs_skipped = ?, docs_failed = ?,"
            " error_message = ? WHERE import_id = ?",
            (int(time.time()), status, prog.imported, prog.skipped, prog.failed,
             error, import_id),
        )

